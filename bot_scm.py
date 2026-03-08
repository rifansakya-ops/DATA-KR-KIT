import time, requests, json, io, os
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

def jalankan_bot():
    print("=== [START] Operasi Bot SCM Izin Prinsip ===")
    # URL Web App yang Anda berikan
    gas_url = "https://script.google.com/macros/s/AKfycbz_RPWn2wVdSpx1UH26r9ZjXrsegreZ261WaXM4woP8ldr_YhqoTjhVus5g022TW4bvYw/exec"
    
    options = uc.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    # Menggunakan Chrome Driver versi stabil di lingkungan GitHub Actions
    driver = uc.Chrome(options=options, version_main=145)
    
    try:
        driver.get("https://scm.nusadaya.net/login")
        wait = WebDriverWait(driver, 25)
        
        # Proses Login Otomatis
        email_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @placeholder='Email atau NIP']")))
        email_input.send_keys(os.environ.get('EMAIL_SCM'))
        driver.find_element(By.XPATH, "//input[@type='password']").send_keys(os.environ.get('PASS_SCM'))
        driver.find_element(By.XPATH, "//button[contains(text(), 'Log in')]").click()
        
        print("Login berhasil, menunggu dashboard...")
        time.sleep(15)
        
        # URL Export yang kita dapatkan dari hasil Fetch/Inspect
        export_url = "https://scm.nusadaya.net/izin-prinsip/export"
        
        # Ambil Cookie/Sesi hasil login untuk proses download
        session_cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        
        # Eksekusi download file Excel
        response_dl = requests.get(export_url, cookies=session_cookies)
        
        if response_dl.status_code == 200:
            print("Download Berhasil. Memproses File Excel...")
            wb = load_workbook(filename=io.BytesIO(response_dl.content), data_only=False)
            
            # Memproses semua tab (sheet) yang ada di dalam file Excel tersebut
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"Membaca data dari tab: {sheet_name}")
                
                rows_data = []
                for row in ws.iter_rows(values_only=False):
                    current_row = []
                    for cell in row:
                        if cell.hyperlink:
                            # Jika sel berisi link, kirim teks + URL-nya
                            current_row.append(f"{cell.value} {cell.hyperlink.target}")
                        else:
                            current_row.append(cell.value if cell.value is not None else "")
                    rows_data.append(current_row)

                # Kirim data ke Google Sheets melalui Apps Script
                if len(rows_data) > 0:
                    res = requests.post(gas_url, data=json.dumps({"rows": rows_data}), headers={'Content-Type': 'application/json'})
                    print(f"Respon [{sheet_name}]: {res.text}")
            
    except Exception as e:
        print(f"Terjadi Kesalahan: {e}")
    finally:
        driver.quit()
        print("=== [FINISH] Operasi Selesai ===")

if __name__ == "__main__":
    jalankan_bot()
